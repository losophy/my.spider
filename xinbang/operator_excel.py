from openpyxl import load_workbook

def ReadAccounts(filename,sheet):
    workbook = load_workbook(filename = filename)
    table = workbook[sheet]

    # 读取第一行属性名
    field_names = []
    for col in range(1, table.max_column+1):
        field_names.append(table.cell(1, col).value)

    initRow = 3
    accounts = []
    for row in range(initRow, table.max_row+1):
        for col in range(1, table.max_column+1):
            field_name = field_names[col-1]
            value = table.cell(row, col).value
            if field_name == "account":
                accounts.append(value)

    return accounts

def WriteExcel(filename,sheet):
    workbook = load_workbook(filename = filename)
    table = workbook[sheet]

    table["E1"] = 'LJK5679842'  
    table['F1'] = 'LKJI66666666666666'  
    workbook.save(filename)  

def ClassifyPublicAccount(filename,sheet):
    accounts = ReadAccounts(filename,sheet)
    for account in accounts:
        print(account)

    # WriteExcel(filename,sheet)
