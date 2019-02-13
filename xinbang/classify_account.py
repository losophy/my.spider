from openpyxl import load_workbook
from selenium import webdriver
from requests import Session
from time import sleep

field_names = []
accounts =[]
classifications = []

def ReadAccounts(filename,sheet):
    workbook = load_workbook(filename = filename)
    table = workbook[sheet]

    # 读取第一行属性名
    for col in range(1, table.max_column+1):
        field_names.append(table.cell(1, col).value)

    initRow = 3
    for row in range(initRow, table.max_row+1):
        for col in range(1, table.max_column+1):
            field_name = field_names[col-1]
            value = table.cell(row, col).value
            if field_name == "account":
                accounts.append(value)

def WriteExcel(filename,sheet):
    workbook = load_workbook(filename = filename)
    table = workbook[sheet]

    initRow = 3
    i = 0
    for row in range(initRow, table.max_row+1):
        for col in range(1, table.max_column+1):
            field_name = field_names[col-1]
            if field_name == "classification":
                table.cell(row, col).value = classifications[i]  
                i = i + 1

    workbook.save(filename)  

def Login():
    chromePath = r'C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe'
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')

    driver = webdriver.Chrome(executable_path=chromePath,chrome_options=chrome_options) 

    loginUrl = 'https://www.newrank.cn/public/login/login.html?back=https%3A//www.newrank.cn/'
    driver.get(loginUrl)

    driver.find_element_by_xpath('/html/body/div[2]/div/div[1]/div[1]/div[2]').click()
    driver.find_element_by_xpath('//*[@id="account_input"]').send_keys('nr_953c4lm34') 
    driver.find_element_by_xpath('//*[@id="password_input"]').send_keys('xinbang123456')
    sleep(1)

    driver.find_element_by_xpath('//*[@id="pwd_confirm"]').click()
    sleep(1)

    return driver

def Classify(driver):
    for account in accounts:
        loginUrl = 'https://www.newrank.cn/public/info/detail.html?account='+account
        driver.get(loginUrl)

        text = driver.find_element_by_xpath('//*[@id="info_detail_head_classify_type"]').text
        classifications.append(text)

    driver.close()

def ClassifyPublicAccount(filename,sheet):
    ReadAccounts(filename,sheet)
    driver = Login()
    Classify(driver)

    # for classification in classifications:
    #     print(classification)

    WriteExcel(filename,sheet)

ClassifyPublicAccount(u"公众号.xlsx","Sheet1")
